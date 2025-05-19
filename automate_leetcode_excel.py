#!/usr/bin/env python3
import os
import re
import json
import argparse
from datetime import datetime
from typing import Optional

import openpyxl
import requests
from dotenv import load_dotenv

# -----------------------------------------------------------------------------
# Configuration helpers
# -----------------------------------------------------------------------------
load_dotenv()

GEMINI_ENDPOINT = (
    "https://generativelanguage.googleapis.com/v1beta/models/"
    "gemini-2.0-flash:generateContent"
)

# -----------------------------------------------------------------------------
# Gemini interaction
# -----------------------------------------------------------------------------

def call_gemini(prompt: str, api_key: Optional[str] = None) -> str:
    """Send *prompt* to the Gemini REST endpoint and return the first reply."""
    api_key = api_key or os.getenv("GEMINI_API_KEY")
    if not api_key:
        raise RuntimeError(
            "Gemini API key not supplied (argument or GEMINI_API_KEY env variable)"
        )

    url = f"{GEMINI_ENDPOINT}?key={api_key}"
    headers = {"Content-Type": "application/json"}
    payload = {"contents": [{"parts": [{"text": prompt}]}]}

    resp = requests.post(url, headers=headers, json=payload, timeout=60)
    resp.raise_for_status()
    data = resp.json()

    try:
        return data["candidates"][0]["content"]["parts"][0]["text"].strip()
    except (KeyError, IndexError, TypeError) as exc:
        raise ValueError(f"Unexpected Gemini response schema: {data}") from exc


def extract_fields(source_code: str, category: str, api_key: Optional[str]) -> dict:
    """Ask Gemini to analyse the *source_code* and return a dict of fields.

    The *technique* field must be 2‑3 well‑formed sentences that explicitly
    describe the key algorithmic trick(s) and why they achieve the stated
    complexity.
    """

    prompt = (
        "You are an expert competitive‑programming reviewer. Examine the "
        "following C++ solution to a LeetCode problem and output a JSON "
        "object with **exactly** these keys (in any order):\n\n"
        "  • problem_number – integer or string (leave empty if unknown)\n"
        "  • title            – title of the LeetCode problem\n"
        "  • difficulty       – Easy/Medium/Hard (best guess)\n"
        "  • space_complexity – Big‑O expression (worst‑case)\n"
        "  • time_complexity  – Big‑O expression (worst‑case)\n"
        "  • runtime          – representative runtime number + units if known, "
        "otherwise empty\n"
        "  • key_concept      – 1‑line primary CS concept (e.g. Hash Map)\n"
        "  • technique        – **2‑3 full sentences** highlighting the major "
        "algorithmic idea/tricks used in this user's solution\n\n"
        "If anything is truly uninferrable, output an empty string.\n"
        "Return **only** the JSON (no markdown, no commentary).\n\n"
        f"Category: {category}\n\nC++ source:\n```cpp\n{source_code}\n```"
    )

    content = call_gemini(prompt, api_key)

    # Strip fenced code block if the model returned one
    if content.startswith("```"):
        match = re.search(r"```(?:json)?\s*([\s\S]*?)\s*```", content)
        content = match.group(1).strip() if match else content

    try:
        return json.loads(content)
    except json.JSONDecodeError as exc:
        raise ValueError(f"Gemini did not return valid JSON:\n{content}") from exc

# -----------------------------------------------------------------------------
# Local helpers
# -----------------------------------------------------------------------------

def read_source(file_path: str) -> str:
    """Return the *entire* file contents as a UTF‑8 string."""
    with open(file_path, "r", encoding="utf-8") as f:
        return f.read()


def append_to_excel(excel_path: str, row: list) -> None:
    """Append *row* to *excel_path* (create workbook + header if needed)."""
    excel_path = os.path.expanduser(excel_path)

    if not os.path.exists(excel_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(
            [
                "Date",
                "Day",
                "Problem #",
                "Title",
                "Difficulty",
                "Category",
                "Space Complexity",
                "Time Complexity",
                "Runtime",
                "Key Concept",
                "Technique",
                "Solution Link",
            ]
        )
    else:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

    ws.append(["" if v is None else str(v) for v in row])
    wb.save(excel_path)


def check_if_problem_exists(excel_path: str, problem_title: str) -> bool:
    """Return True if *problem_title* already present in column 4."""
    excel_path = os.path.expanduser(excel_path)
    if not os.path.exists(excel_path):
        return False

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, max_col=4, values_only=True):
        if row[3] and row[3].strip() == problem_title:
            return True
    return False

# -----------------------------------------------------------------------------
# CLI entry‑point
# -----------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Populate an Excel log of LeetCode solutions via Gemini analysis"
    )
    parser.add_argument("files", nargs="+", help="Paths to C++ source files")
    parser.add_argument("--excel", required=True, help="Target Excel .xlsx path")
    parser.add_argument("--day", type=int, required=True, help="Day counter")
    parser.add_argument("--api-key", help="Gemini API key (overrides env)")
    parser.add_argument("--date", help="Custom date in MM/DD/YYYY format (overrides today's date)")
    args = parser.parse_args()

    today = args.date or datetime.today().strftime("%m/%d/%Y")

    for file_path in args.files:
        category = os.path.basename(os.path.dirname(file_path)) or "Uncategorised"
        source_code = read_source(file_path)

        data = extract_fields(source_code, category, args.api_key)

        if check_if_problem_exists(args.excel, data.get("title", "")):
            print(f"⏩  Skipping '{data.get('title')}' (already logged)")
            continue

        append_to_excel(
            args.excel,
            [
                today,
                args.day,
                data.get("problem_number"),
                data.get("title"),
                data.get("difficulty"),
                category,
                data.get("space_complexity"),
                data.get("time_complexity"),
                data.get("runtime"),
                data.get("key_concept"),
                data.get("technique"),
                "",  # manual solution link
            ],
        )
        print(f"✅ Logged '{data.get('title')}' → {os.path.abspath(args.excel)}")


if __name__ == "__main__":
    main()

#Default Values: leet "Arrays & Hashing/valid_anagram.cpp"
#Custom Date and Day: leet --date 05/18/2025 --day 2 "Arrays & Hashing/valid_anagram.cpp"